import openpyxl
import os

ROWS_IN_TEST = 11
ROWS_IN_SOURCE  = 11

workbook1 = openpyxl.load_workbook('C:/Users/Ivgeni/Desktop/PROJECTS/python/excel/test.xlsx')
workbook2 = openpyxl.load_workbook('C:/Users/Ivgeni/Desktop/PROJECTS/python/excel/source.xlsx')
workbook3 = openpyxl.Workbook()

sheet1 = workbook1['Sheet1']
sheet2 = workbook2['Sheet1']
sheet3 = workbook3['Sheet']

counter = 0
row = 2
for i  in range(2,ROWS_IN_TEST):
    for j  in range(2,ROWS_IN_SOURCE):
        if sheet1.cell(row =i, column =1).value == sheet2.cell(row = j, column= 1).value:
            if sheet1.cell(row =i, column =2).value == sheet2.cell(row = j, column= 2).value:
                counter += 1
                print(str(sheet1.cell(row =i, column =1).value) + " " + str(sheet1.cell(row =i, column =2).value) + " " + str(sheet2.cell(row =j, column =3).value) )
                sheet3['A' + str(row)] = sheet1.cell(row =i, column =1).value
                sheet3['B' + str(row)] = sheet1.cell(row =i, column =2).value
                sheet3['C' + str(row)] = sheet2.cell(row =j, column =3).value
                row += 1
            else:
                print("Mismatch in names for value in line " + str(i))

    if counter > 1:
        print("Too many values found for value in line: " + str(i))
    if counter == 0:
        print("No match found for value in line: " + str(i))

    counter = 0

    workbook3.save('C:/Users/Ivgeni/Desktop/PROJECTS/python/excel/output.xlsx')

